"""공급처별 발주 엑셀 생성 (양식은 config.yaml 의 order_sheet.columns 로 정의)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DEFAULT_COLUMNS = [
    {"header": "수취인명", "field": "receiver_name"},
    {"header": "연락처", "field": "receiver_phone"},
    {"header": "우편번호", "field": "receiver_zip"},
    {"header": "주소", "field": "receiver_addr"},
    {"header": "상품명", "field": "supplier_item_name"},
    {"header": "수량", "field": "qty"},
    {"header": "배송메시지", "field": "delivery_message"},
    {"header": "주문번호", "field": "order_id"},
]


def generate_order_sheet(file_path: str | Path, rows: list[dict], columns: list[dict] | None = None) -> Path:
    """rows: 발주 라인(dict) 목록. columns 의 field 값을 뽑아 엑셀로 저장한다."""
    columns = columns or DEFAULT_COLUMNS
    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "발주서"

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col["header"])
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col["field"], ""))

    for col_idx, col in enumerate(columns, start=1):
        values = [str(col["header"])] + [str(r.get(col["field"], "")) for r in rows]
        width = min(40, max(10, max(len(v) for v in values) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(file_path)
    return file_path
