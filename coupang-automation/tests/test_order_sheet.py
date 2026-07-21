from openpyxl import load_workbook

from coupang_auto.order_sheet import generate_order_sheet


def test_generate_order_sheet(tmp_path):
    columns = [
        {"header": "받는분", "field": "receiver_name"},
        {"header": "품목명", "field": "supplier_item_name"},
        {"header": "수량", "field": "qty"},
    ]
    rows = [
        {"receiver_name": "홍길동", "supplier_item_name": "성주참외 2kg", "qty": 2},
        {"receiver_name": "김철수", "supplier_item_name": "하우스감귤 1.5kg", "qty": 1},
    ]
    path = generate_order_sheet(tmp_path / "po.xlsx", rows, columns)

    ws = load_workbook(path).active
    assert [c.value for c in ws[1]] == ["받는분", "품목명", "수량"]
    assert [c.value for c in ws[2]] == ["홍길동", "성주참외 2kg", 2]
    assert [c.value for c in ws[3]] == ["김철수", "하우스감귤 1.5kg", 1]
